"""Masterfile-Zugriff: doc_id -> MMSID (Alma/Swisscovery-Norm-ID).

Die ZBZ-Masterfile (``data/source/masterfile/Masterfile.xlsx``) traegt pro Text in
der Spalte ``MMSID`` die Alma-MMSID, indexiert ueber die Spalte ``ID`` (= doc_id).
Sie ist die Single Source of Truth fuer diese Norm-ID -- damit ist die
Daten-Verfuegbarkeit von O8 (Metadaten aus Alma/MMSID) abgedeckt; der TEI-Header
kann die MMSID als ``<idno type="MMSID">`` fuehren.

Der Loader liest ID + MMSID einmal und cached die Map (read-only).
"""

import functools

from scripts.config import MASTERFILE_PATH


def _norm_id(value) -> str | None:
    """Normalisiert eine Masterfile-ID zu str ohne ``.0`` (10, 10.0, '10' -> '10')."""
    try:
        return str(int(value))
    except (ValueError, TypeError):
        s = str(value).strip()
        return s or None


@functools.lru_cache(maxsize=1)
def mmsid_map() -> dict:
    """{doc_id: mmsid} aus der Masterfile. Leer, wenn Datei/openpyxl/Spalten fehlen."""
    if not MASTERFILE_PATH.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        return {}

    wb = openpyxl.load_workbook(MASTERFILE_PATH, read_only=True, data_only=True)
    ws = wb["Tabelle1"] if "Tabelle1" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    if "ID" not in header or "MMSID" not in header:
        return {}
    i_id, i_mms = header.index("ID"), header.index("MMSID")

    out = {}
    for r in rows[1:]:
        if not r:
            continue
        doc_id = _norm_id(r[i_id])
        mms = r[i_mms]
        if doc_id is None or mms in (None, ""):
            continue
        out[doc_id] = str(mms).strip()
    return out


def mmsid_for(doc_id) -> str | None:
    """MMSID fuer eine doc_id, oder None wenn nicht in der Masterfile."""
    return mmsid_map().get(str(doc_id))
