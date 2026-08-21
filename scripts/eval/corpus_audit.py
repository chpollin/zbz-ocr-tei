"""
Korpus-Audit: leitet alle Korpus-Kennzahlen deterministisch aus den
Primaerquellen ab und macht die Reconciliation-Kaskade explizit.

Hintergrund: Die Zahlen in knowledge/*.md sind LLM-kuratiert und driften --
insbesondere durch Vermischung der Zaehl-Einheiten (Text-Ebene laut Masterfile
vs PDF-/Pipeline-Ebene). Dieses Skript bindet jede Kennzahl an ein Tripel
(Quelle, Einheit, Extraktion) und legt den Trichter 325 -> 289 -> 286 -> 285 offen.

Quellen-Tiers (nach Autoritaet geordnet):
  0  Masterfile.xlsx     ZBZ-bibliografisch, Text-Ebene (Genre, Sprache, Jahr, Seiten)
  1  data/source/pdf/*.pdf    physisch geliefert (PDF-Ebene, Seiten via pypdfium2)
  2  output/ + docs/     Pipeline-Output (deterministische Datei-/Element-Zaehlung)
  3  doc_metadata.json   Gemini-Klassifikation (abgeleitet -- nur Hypothese)
  4  knowledge/*.md      reines Verifikations-Ziel, nie Quelle (-> drift-Check)

Usage:
    python -m scripts.eval.corpus_audit               # JSON + Markdown nach output/
    python -m scripts.eval.corpus_audit --print       # Markdown zusaetzlich auf stdout
    python -m scripts.eval.corpus_audit --json PATH --md PATH
"""

import argparse
import json
import re
import statistics
import subprocess
from collections import Counter
from datetime import UTC, datetime
from pathlib import Path

from scripts.config import (
    DOC_METADATA_PATH,
    IMAGES_DIR,
    LAYOUT_DIR,
    MASTERFILE_PATH,
    MISTRAL_RESULTS_DIR,
    OUTPUT_DIR,
    PAGE_XML_DIR,
    PROJECT_ROOT,
    SCANS_DIR,
    TEI_FINAL_DIR,
)

TOOL_VERSION = "1.0"

# Verified target values (as of project.md 2026-05-27) for the drift check.
# This baseline IS the knowledge claim; update here when the knowledge docs change.
KNOWLEDGE_CLAIMS = {
    "masterfile_texts": 325,   # project.md: Masterfile texts (text level)
    "pdfs": 286,               # project.md: delivered as PDF
    "tei_produced": 285,       # project.md: productive (final TEI)
    "pages_total": 7186,       # project.md: bibliographic pages (Masterfile)
    "pages_processed": 4122,   # project.md: processed pages (OCR md, volatile on re-OCR)
}


def _id_sort_key(x):
    return int(x) if x and x.isdigit() else 0


def _stats(values):
    """Min/Median/Max/Summe fuer eine Zahlenliste (None-sicher)."""
    nums = [v for v in values if isinstance(v, (int, float))]
    if not nums:
        return {"n": 0, "sum": 0, "median": None, "min": None, "max": None}
    return {
        "n": len(nums),
        "sum": int(sum(nums)),
        "median": float(statistics.median(nums)),
        "min": min(nums),
        "max": max(nums),
    }


def _norm_id(value):
    """Normalisiert eine Dokument-ID zu einem String ('10' statt 10.0)."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s


def _git_sha():
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=PROJECT_ROOT, stderr=subprocess.DEVNULL,
        ).decode().strip()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Tier 0 -- Masterfile (ZBZ-bibliografisch, Text-Ebene)
# ---------------------------------------------------------------------------

def audit_masterfile():
    result = {
        "source": "data/source/masterfile/Masterfile.xlsx",
        "unit": "Texte (Masterfile-Zeilen mit ID)",
        "available": False,
    }
    if not MASTERFILE_PATH.exists():
        result["error"] = "Masterfile nicht gefunden"
        return result
    try:
        import openpyxl
    except ImportError:
        result["error"] = "openpyxl nicht installiert"
        return result

    wb = openpyxl.load_workbook(MASTERFILE_PATH, read_only=True, data_only=True)
    ws = wb["Tabelle1"] if "Tabelle1" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col = {
        name: header.index(name)
        for name in ("ID", "PublForm", "Jahr", "Anzahl Seiten", "Sprache", "digitalisiert")
        if name in header
    }
    if "ID" not in col:
        result["error"] = "Spalte 'ID' im Header nicht gefunden"
        return result

    recs = [r for r in rows[1:] if r and r[col["ID"]] not in (None, "")]
    ids = {_norm_id(r[col["ID"]]) for r in recs}
    ids.discard(None)

    def vals(name):
        i = col.get(name)
        return [r[i] for r in recs] if i is not None else []

    publform = Counter((str(v).strip() if v else "(leer)") for v in vals("PublForm"))
    sprache = Counter((str(v).strip() if v else "(leer)") for v in vals("Sprache"))
    digital = Counter((str(v).strip().lower() if v else "(leer)") for v in vals("digitalisiert"))

    dig_ja_ids = set()
    if "digitalisiert" in col:
        for r in recs:
            if str(r[col["digitalisiert"]]).strip().lower() == "ja":
                dig_ja_ids.add(_norm_id(r[col["ID"]]))
    dig_ja_ids.discard(None)

    jahre = []
    for v in vals("Jahr"):
        m = re.search(r"(?:19|20)\d{2}", str(v or ""))
        if m:
            jahre.append(int(m.group()))

    pages = []
    for v in vals("Anzahl Seiten"):
        try:
            pages.append(int(float(str(v).replace(",", "."))))
        except (ValueError, TypeError):
            pass

    # Pro-ID-Map, damit nachgelagert auf eine Teilmenge (z.B. die gelieferten
    # PDFs) gefiltert werden kann, ohne die Masterfile erneut zu lesen.
    def _cell(r, name):
        i = col.get(name)
        v = r[i] if i is not None else None
        return str(v).strip() if v not in (None, "") else "(leer)"

    def _year(r):
        i = col.get("Jahr")
        m = re.search(r"(?:19|20)\d{2}", str(r[i] or "")) if i is not None else None
        return int(m.group()) if m else None

    by_id = {}
    for r in recs:
        rid = _norm_id(r[col["ID"]])
        if rid is not None:
            by_id[rid] = {"publform": _cell(r, "PublForm"),
                          "sprache": _cell(r, "Sprache"),
                          "jahr": _year(r)}

    result.update({
        "available": True,
        "texts": len(recs),
        "by_id": by_id,
        "unique_ids": len(ids),
        "ids": sorted(ids, key=_id_sort_key),
        "digitalisiert": dict(digital.most_common()),
        "digitalisiert_ja_ids": sorted(dig_ja_ids, key=_id_sort_key),
        "publform": dict(publform.most_common()),
        "sprache": dict(sprache.most_common()),
        "jahr": {
            "min": min(jahre) if jahre else None,
            "max": max(jahre) if jahre else None,
            "count_1970_1989": sum(1 for j in jahre if 1970 <= j <= 1989),
        },
        "anzahl_seiten": _stats(pages),
    })
    return result


# ---------------------------------------------------------------------------
# Tier 1 -- gelieferte PDFs (physisch)
# ---------------------------------------------------------------------------

def _pdf_page_count(pdf_path):
    """Physische Seitenzahl: pypdfium2 (exakt), Fallback Regex."""
    try:
        import pypdfium2 as pdfium
        doc = pdfium.PdfDocument(str(pdf_path))
        n = len(doc)
        doc.close()
        return n, "pypdfium2"
    except Exception:
        try:
            content = pdf_path.read_bytes()
            n = len(re.findall(rb"/Type\s*/Page(?!\s*s)", content))
            return (n if n > 0 else None), "regex"
        except Exception:
            return None, "error"


def audit_pdfs():
    pdfs = sorted(SCANS_DIR.glob("*.pdf")) if SCANS_DIR.exists() else []
    ids = {_norm_id(p.stem) for p in pdfs}
    ids.discard(None)
    page_counts = []
    methods = Counter()
    for p in pdfs:
        n, m = _pdf_page_count(p)
        methods[m] += 1
        if n is not None:
            page_counts.append(n)
    return {
        "source": "data/source/pdf/*.pdf",
        "unit": "PDFs (physisch geliefert)",
        "pdfs": len(pdfs),
        "ids": sorted(ids, key=_id_sort_key),
        "physical_pages": _stats(page_counts),
        "page_count_method": dict(methods),
    }


# ---------------------------------------------------------------------------
# Tier 2 -- Pipeline-Output (deterministisch)
# ---------------------------------------------------------------------------

def _doc_from_page_file(name):
    """'1330_p2.md' -> '1330'"""
    return name.rsplit(".", 1)[0].split("_p")[0]


def _count_pb(path):
    try:
        return len(re.findall(r"<pb[\s/>]", path.read_text(encoding="utf-8")))
    except Exception:
        return 0


def audit_pipeline():
    md_files = list(MISTRAL_RESULTS_DIR.glob("*.md")) if MISTRAL_RESULTS_DIR.exists() else []
    ocr_docs = {_doc_from_page_file(p.name) for p in md_files}
    finals = sorted(TEI_FINAL_DIR.glob("*_final.xml")) if TEI_FINAL_DIR.exists() else []
    final_ids = {_norm_id(p.name.replace("_final.xml", "")) for p in finals}
    final_ids.discard(None)
    manifests = list(TEI_FINAL_DIR.glob("*_manifest.json")) if TEI_FINAL_DIR.exists() else []
    image_dirs = [d for d in IMAGES_DIR.iterdir() if d.is_dir()] if IMAGES_DIR.exists() else []
    layout_dirs = [d for d in LAYOUT_DIR.iterdir() if d.is_dir()] if LAYOUT_DIR.exists() else []
    pagexml_dirs = [d for d in PAGE_XML_DIR.iterdir() if d.is_dir()] if PAGE_XML_DIR.exists() else []
    return {
        "source": "output/ + docs/",
        "unit": "Pipeline-Artefakte",
        "ocr_md_pages": len(md_files),
        "ocr_docs": len(ocr_docs),
        "tei_final_docs": len(finals),
        "tei_final_ids": sorted(final_ids, key=_id_sort_key),
        "tei_pb_total": sum(_count_pb(p) for p in finals),
        "manifests": len(manifests),
        "image_dirs": len(image_dirs),
        "layout_dirs": len(layout_dirs),
        "page_xml_dirs": len(pagexml_dirs),
    }


# ---------------------------------------------------------------------------
# Tier 3 -- doc_metadata.json (Gemini-Klassifikation, abgeleitet)
# ---------------------------------------------------------------------------

def audit_doc_metadata():
    if not DOC_METADATA_PATH.exists():
        return {"available": False, "error": "doc_metadata.json nicht gefunden"}
    data = json.loads(DOC_METADATA_PATH.read_text(encoding="utf-8"))
    docs = data.get("documents", {})
    publform = Counter(v.get("pub_form") or "(leer)" for v in docs.values())
    language = Counter(v.get("language") or "(leer)" for v in docs.values())
    return {
        "available": True,
        "source": "data/doc_metadata.json",
        "unit": "PDFs (Gemini-Klassifikation)",
        "generated": data.get("generated"),
        "model": data.get("model"),
        "total_docs_field": data.get("total_docs"),
        "entries": len(docs),
        "page_count": _stats([v.get("page_count") for v in docs.values()]),
        "pub_form": dict(publform.most_common()),
        "language": dict(language.most_common()),
    }


# ---------------------------------------------------------------------------
# Reconciliation + Drift
# ---------------------------------------------------------------------------

def reconcile(t0, t1, t2, t3):
    mf_ids = set(t0.get("ids", []))
    dig_ids = set(t0.get("digitalisiert_ja_ids", []))
    pdf_ids = set(t1.get("ids", []))
    tei_ids = set(t2.get("tei_final_ids", []))
    return {
        "funnel": [
            {"stage": "Masterfile-Texte", "n": t0.get("texts"), "source": "Tier 0"},
            {"stage": "davon digitalisiert=ja", "n": t0.get("digitalisiert", {}).get("ja"), "source": "Tier 0"},
            {"stage": "davon als PDF geliefert", "n": t1.get("pdfs"), "source": "Tier 1"},
            {"stage": "davon mit finalem TEI", "n": t2.get("tei_final_docs"), "source": "Tier 2"},
        ],
        "scans_not_in_masterfile": sorted(pdf_ids - mf_ids, key=_id_sort_key),
        "digitalisiert_not_delivered": sorted(dig_ids - pdf_ids, key=_id_sort_key),
        "scans_without_final_tei": sorted(pdf_ids - tei_ids, key=_id_sort_key),
        "page_counts_by_source": {
            "masterfile_bibliographic": t0.get("anzahl_seiten", {}).get("sum"),
            "pdf_physical": t1.get("physical_pages", {}).get("sum"),
            "ocr_md": t2.get("ocr_md_pages"),
            "tei_pb": t2.get("tei_pb_total"),
            "gemini_page_count_field": t3.get("page_count", {}).get("sum") if t3.get("available") else None,
        },
    }


def delivered_distribution(t0, pdf_ids):
    """Verteilungen (Genre/Sprache/Jahr) ueber die GELIEFERTEN Dokumente.

    Schnittmenge Masterfile-Metadaten und gelieferte PDFs (n=286), im Gegensatz
    zu den Katalog-Verteilungen in Tier 0 (n=325). Das ist die Sicht, mit der
    projektseitig gearbeitet wird ("nur gelieferte Daten").
    """
    by_id = t0.get("by_id", {})
    pdf_ids = set(pdf_ids)
    delivered = [by_id[i] for i in pdf_ids if i in by_id]
    jahre = [d["jahr"] for d in delivered if d["jahr"]]
    return {
        "n": len(delivered),
        "not_in_masterfile": sorted(pdf_ids - set(by_id), key=_id_sort_key),
        "publform": dict(Counter(d["publform"] for d in delivered).most_common()),
        "sprache": dict(Counter(d["sprache"] for d in delivered).most_common()),
        "jahr": {
            "min": min(jahre) if jahre else None,
            "max": max(jahre) if jahre else None,
            "count_1970_1989": sum(1 for j in jahre if 1970 <= j <= 1989),
        },
    }


def drift_check(t0, t1, t2, t3):
    computed = {
        "masterfile_texts": t0.get("texts"),
        "pdfs": t1.get("pdfs"),
        "tei_produced": t2.get("tei_final_docs"),
        "pages_total": t0.get("anzahl_seiten", {}).get("sum"),
        "pages_processed": t2.get("ocr_md_pages"),
    }
    notes = {
        "masterfile_texts": "Text-Ebene (Masterfile-Zeilen mit ID), nicht der digitalisiert-Zaehler (289)",
        "pages_total": "bibliografische Summe (Masterfile, n=325), nicht physisch/verarbeitet",
        "pages_processed": "OCR-Markdown-Seiten; TEI-<pb> = 4.115 (Assembly-Artefakt)",
    }
    out = []
    for key, claimed in KNOWLEDGE_CLAIMS.items():
        comp = computed.get(key)
        status = "OK" if comp == claimed else "DRIFT"
        out.append({
            "metric": key,
            "claimed_knowledge": claimed,
            "computed": comp,
            "status": status,
            "note": notes.get(key, ""),
        })
    return out


# ---------------------------------------------------------------------------
# Report-Assembly + Markdown
# ---------------------------------------------------------------------------

def build_report():
    t0 = audit_masterfile()
    t1 = audit_pdfs()
    t2 = audit_pipeline()
    t3 = audit_doc_metadata()
    return {
        "meta": {
            "tool_version": TOOL_VERSION,
            "generated_at": datetime.now(UTC).isoformat(),
            "git_sha": _git_sha(),
        },
        "tier0_masterfile": t0,
        "tier1_pdfs": t1,
        "tier2_pipeline": t2,
        "tier3_doc_metadata": t3,
        "reconciliation": reconcile(t0, t1, t2, t3),
        "delivered_distribution": delivered_distribution(t0, t1.get("ids", [])),
        "drift_vs_knowledge": drift_check(t0, t1, t2, t3),
    }


def _dist_table(title, dist):
    lines = [f"**{title}**", "", "| Kategorie | n |", "|---|---|"]
    for k, v in dist.items():
        lines.append(f"| {k} | {v} |")
    return "\n".join(lines)


def render_markdown(rep):
    t0, t1, t2, t3 = (rep["tier0_masterfile"], rep["tier1_pdfs"],
                      rep["tier2_pipeline"], rep["tier3_doc_metadata"])
    rec = rep["reconciliation"]
    md = []
    md.append("# Korpus-Audit")
    md.append("")
    md.append(f"Generiert: {rep['meta']['generated_at']} | Commit: {rep['meta']['git_sha']} "
              f"| Tool v{rep['meta']['tool_version']}")
    md.append("")
    md.append("Jede Zahl ist an (Quelle, Einheit, Extraktion) gebunden. "
              "Tier 0 = Masterfile (Text-Ebene), Tier 1 = PDFs (physisch), "
              "Tier 2 = Pipeline-Output, Tier 3 = Gemini (abgeleitet).")
    md.append("")

    md.append("## Reconciliation-Trichter")
    md.append("")
    md.append("| Stufe | n | Quelle |")
    md.append("|---|---|---|")
    for s in rec["funnel"]:
        md.append(f"| {s['stage']} | {s['n']} | {s['source']} |")
    md.append("")
    md.append(f"- in Scans, nicht in Masterfile: {rec['scans_not_in_masterfile'] or 'keine'}")
    md.append(f"- digitalisiert, aber nicht als PDF geliefert: {rec['digitalisiert_not_delivered'] or 'keine'}")
    md.append(f"- PDF ohne finales TEI: {rec['scans_without_final_tei'] or 'keine'}")
    md.append("")

    md.append("## Seitenzahlen nach Quelle (verschiedene Dinge!)")
    md.append("")
    md.append("| Definition | Wert | Einheit |")
    md.append("|---|---|---|")
    pcs = rec["page_counts_by_source"]
    md.append(f"| bibliografisch (Masterfile) | {pcs['masterfile_bibliographic']} | Texte (n=325) |")
    md.append(f"| physisch (PDFs) | {pcs['pdf_physical']} | PDFs (n=286) |")
    md.append(f"| verarbeitet (OCR-Markdown) | {pcs['ocr_md']} | Pipeline-Seiten |")
    md.append(f"| verarbeitet (TEI <pb>) | {pcs['tei_pb']} | TEI-Seiten |")
    md.append(f"| Gemini page_count-Summe | {pcs['gemini_page_count_field']} | abgeleitet (nicht echt) |")
    md.append("")

    dd = rep["delivered_distribution"]
    md.append(f"## Gelieferte Dokumente (n={dd['n']}) -- Verteilungen")
    md.append("")
    md.append("Schnittmenge Masterfile-Metadaten und gelieferte PDFs -- die projektseitig "
              "massgebliche Sicht (im Gegensatz zur Katalog-Ebene n=325 in Tier 0).")
    md.append(f"Jahr {dd['jahr']['min']}-{dd['jahr']['max']}, "
              f"{dd['jahr']['count_1970_1989']} in 1970-1989.")
    if dd["not_in_masterfile"]:
        md.append(f"Geliefert, aber nicht in Masterfile: {dd['not_in_masterfile']}")
    md.append("")
    md.append(_dist_table("Genre / PublForm (geliefert)", dd["publform"]))
    md.append("")
    md.append(_dist_table("Sprache (geliefert)", dd["sprache"]))
    md.append("")

    md.append("## Drift gegen Knowledge-Base")
    md.append("")
    md.append("| Metrik | Knowledge | Berechnet | Status | Hinweis |")
    md.append("|---|---|---|---|---|")
    for d in rep["drift_vs_knowledge"]:
        md.append(f"| {d['metric']} | {d['claimed_knowledge']} | {d['computed']} "
                  f"| {d['status']} | {d['note']} |")
    md.append("")

    if t0.get("available"):
        md.append("## Tier 0 -- Masterfile (Text-Ebene, autoritativ fuer Bibliografie)")
        md.append("")
        md.append(f"Texte: {t0['texts']} | digitalisiert: {t0['digitalisiert']} | "
                  f"Jahr {t0['jahr']['min']}-{t0['jahr']['max']}, "
                  f"{t0['jahr']['count_1970_1989']} in 1970-1989 | "
                  f"Seiten (biblio): Summe {t0['anzahl_seiten']['sum']}, "
                  f"Median {t0['anzahl_seiten']['median']}, Max {t0['anzahl_seiten']['max']}")
        md.append("")
        md.append(_dist_table("PublForm (Masterfile)", t0["publform"]))
        md.append("")
        md.append(_dist_table("Sprache (Masterfile)", t0["sprache"]))
        md.append("")

    if t3.get("available"):
        md.append("## Tier 3 -- Gemini-Klassifikation (PDF-Ebene, abgeleitet)")
        md.append("")
        md.append(f"Modell: {t3['model']} | Eintraege: {t3['entries']} | "
                  f"page_count: Summe {t3['page_count']['sum']}, "
                  f"Median {t3['page_count']['median']}, Max {t3['page_count']['max']}")
        md.append("")
        md.append(_dist_table("pub_form (Gemini)", t3["pub_form"]))
        md.append("")
        md.append(_dist_table("language (Gemini)", t3["language"]))
        md.append("")

    return "\n".join(md)


def main():
    ap = argparse.ArgumentParser(description="Korpus-Audit aus den Primaerquellen")
    ap.add_argument("--json", type=Path, default=OUTPUT_DIR / "corpus_audit.json",
                    help="Pfad fuer JSON-Output")
    ap.add_argument("--md", type=Path, default=OUTPUT_DIR / "corpus_audit.md",
                    help="Pfad fuer Markdown-Output")
    ap.add_argument("--print", dest="do_print", action="store_true",
                    help="Markdown zusaetzlich auf stdout")
    args = ap.parse_args()

    rep = build_report()
    args.json.parent.mkdir(parents=True, exist_ok=True)
    args.json.write_text(json.dumps(rep, ensure_ascii=False, indent=2), encoding="utf-8")
    md = render_markdown(rep)
    args.md.write_text(md, encoding="utf-8")

    # ASCII-sichere Konsolen-Zusammenfassung
    rec = rep["reconciliation"]
    print("Korpus-Audit geschrieben:")
    print(f"  JSON: {args.json}")
    print(f"  MD:   {args.md}")
    print("Trichter: " + " -> ".join(f"{s['n']}" for s in rec["funnel"]))
    print(f"PDF ohne TEI: {rec['scans_without_final_tei'] or 'keine'}")
    drift = [d for d in rep["drift_vs_knowledge"] if d["status"] == "DRIFT"]
    print(f"Drift gegen Knowledge-Base: {len(drift)} von {len(rep['drift_vs_knowledge'])} Metriken")
    for d in drift:
        print(f"  - {d['metric']}: Knowledge={d['claimed_knowledge']} berechnet={d['computed']}")
    if args.do_print:
        print("\n" + md)


if __name__ == "__main__":
    main()
